#!/usr/bin/python
# -*- encoding: utf-8 -*-
import sys
reload(sys)    # re-enable sys.setdefaultencoding()
sys.setdefaultencoding('utf-8')
import os
from flask import Flask, request, redirect, url_for, g
from werkzeug import secure_filename
from settings.config import *
from openpyxl import Workbook
from openpyxl import load_workbook
import random, string
from pprint import pprint

########################
# EXCEL CREATION CLASS #
########################

class Excel:

  def __init__(self, 
      upload_folder = None,
      allowed_extensions = None 
      ):
    if not upload_folder:
      upload_folder = app.config['UPLOAD_FOLDER'] 
    if not allowed_extensions:
      allowed_extensions = set(['xlsx'])
    self.upload_folder = upload_folder
    self.allowed_extensions = allowed_extensions
    if not os.path.exists(self.upload_folder):
      os.makedirs(self.upload_folder)
    return

  def random_string(self, length = 10):
     return ''.join(random.choice(string.lowercase) for i in range(length))

  def random_filename(self, folder = None, length = 10, extension = '.xlsx'):
    if not folder:
      folder = self.upload_folder
    random_string = self.random_string(length)
    random_filename = random_string + extension
    check = os.path.isfile(folder + '/' + random_filename ) 
    while os.path.isfile(folder + '/' + random_filename ):
      random_string = self.random_string()
      random_filename = random_string + extension
    return random_filename

  # Generate an example xlsx file from the defaults edit columns of a model schema
  def generate_example(self, model = None, columns = None, filename = None):
    wb = Workbook()
    ws = wb.active
    if not columns and model:
      try:
        ws.append(model['defaults']['import'])
      except Exception as exc:
        return
    else:
      ws.append(columns)
    if not os.path.exists(BASE_DIR + '/static/assets/examples'):
      os.makedirs(BASE_DIR + '/static/assets/examples')
      
    if filename:
      filename = BASE_DIR + '/static/assets/examples/' + filename
    elif model:
      filename = BASE_DIR + '/static/assets/examples/' +  model['table'] + ".xlsx"
    wb.save(filename)
    return True

  def export_model(self, model, columns = []):
    if not os.path.exists(self.upload_folder + "/export/"):
      os.makedirs(self.upload_folder + "/export/")
    filename = self.random_string(24)
    xlsx_file = self.upload_folder + "/export/" +  filename + ".xlsx"
    wb = Workbook()
    ws = wb.active

    # set columns and add heading row #
    if type(columns) is list and len(columns):
      model_columns = []
      for model_column in model['columns']:
        model_columns.append(model_column[0])
      for column in columns:
        if column not in model_columns:
          print "Error: column not found " + column
          return False
    else:
      for column in model['columns']:
        columns.append(column[0])

    ws.append(columns)

    # append every row #
    db = g.db
    data = db.read("SELECT " + ', '.join(columns) + " FROM " + model['table'])
    test_data = db.foreign_replace(model, data)
    for item in data:
      row = []
      for column in columns:
        row.append(item[column])
      ws.append(row)
    try:
      wb.save(xlsx_file)
      with open(xlsx_file, 'r') as content_file:
        content = content_file.read()
      os.remove(xlsx_file)
    except Exception as exc:
      print "Error: ", exc
      return False

    return content

  def file_upload(self, file = None, custom_filename = None):
    # Upload File From Form Post
    if request.method == 'POST':
      try:
        file = request.files['file']
        filename = secure_filename(file.filename)
        check = filename.rsplit('.', 1)[1] in self.allowed_extensions
      except Exception as e:
        print e
        return False
      if file and check:
        try:
          if custom_filename:
            filename = custom_filename
          file.save(os.path.join(self.upload_folder, filename))
          return filename
          #return redirect(url_for('uploaded_file',filename=filename))
        except Exception as e:
          print e
          return False
      else:
        return False

  def read(self, filename, format = 'xlsx', remove_after = False):
    try:
      wb = load_workbook(os.path.join(self.upload_folder, filename), use_iterators=True, data_only = True)
    except Exception as e:
      wb = load_workbook(os.path.join(self.upload_folder, filename), data_only = True)
    
    wss = wb.get_sheet_names()
    ws = wb.active

    # Convert excel worksheet tupple into a list
    data = []
    for row in ws.iter_rows():
      item = []
      for cell in row:
        value = unicode(cell.value)
        item.append(value)
      data.append(item)  

    # Create dictionary with results
    thead = data[0]
    del data[0]
    data = {
        'thead' : thead,
        'tbody' : data
        }

    if remove_after:
        os.remove(self.upload_folder + '/' + filename)

    return data

  def import_file(self, model, filename = None, columns = None, strict_rules = False):
    # 1. upload the file: http://flask.pocoo.org/docs/0.10/patterns/fileuploads/
    # 2. read the file and check columns
    # 3. convert, loop and write the data in the database
    if not filename:
      filename = self.file_upload()
    
    if not columns:
      try:
        columns = model['defaults']['import']
      except Exception as exc:
        print "Error columns: ", exc
        return False
       
    # Start processing file
    if filename:
      wb = load_workbook(os.path.join(self.upload_folder, filename), use_iterators=False, data_only = True)
      wss = wb.get_sheet_names()
      ws = wb.active

      # Convert excel worksheet tupple into a list
      data = []
      for row in ws.rows:
        item = []
        for cell in row:
          #value = unicode(cell.value).encode(encoding='UTF-8',errors='strict')
          value = unicode(cell.value)
          item.append(value)
        data.append(item)  

      # Store columns names in a variable and remove it from data results
      col_names = data[0]
      del data[0]
      
      # Check if imported column names exist in authorized column names
      for col in col_names:
        if col not in columns:
          print "Invalid column in imported file: " + str(col)
          return False

      if strict_rules:
        # Check if all allowed columns names exists in imported file
        for column in columns:
          if column not in data[0]:
            print "Missing column in imported file: " + column
            return False
      
      # Add data to database
      placeholders = []
      for item in col_names:
        placeholders.append("?")
      placeholders = "(" + ", ".join(placeholders) + ")"

      col_names = ", ".join(col_names)
        
      try:
        db = model['db']
        table = model['table']
        query = "INSERT INTO " + table + " (" + col_names + ") VALUES " + placeholders
        db.mod(query ,data)
      except Exception as exc:
        print "Model DB Insertion Error: " + str(exc)
        return False
      return

